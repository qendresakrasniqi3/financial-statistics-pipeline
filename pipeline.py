# =============================================================================
# pipeline.py
# Euro Area Payments Statistics Pipeline — Germany (DE)
#
# DISCLAIMER:
#   This pipeline uses synthetic data generated to approximate real payment
#   statistics for Germany as published online. Volumes and values are
#   calibrated to realistic orders of magnitude but numbers may slightly vary
#   from official published figures. This dataset is intended for
#   demonstration purposes only.
#
# 1. INGEST      — Load quarterly payment datasets, schema validation,
#                  statistical summary
# 2. VALIDATE    — Business rules, referential integrity,
#                  data quality assessment
# 3. TRANSFORM   — Data joins, derived variables,
#                  statistical indicators, QoQ and YoY calculations
# 4. ANALYSE     — Country-level trends, outlier detection
#                  (Z-score, IQR), quarter-on-quarter spike detection
# 5. REPORT      — Germany payments statistical summary tables,
#                  dashboard-ready outputs
# 6. SQL         — Business queries using sqlite3
# 7. TEST        — Unit tests, pipeline validation,
#                  reconciliation checks, output consistency checks
# =============================================================================

import pandas as pd
import numpy as np
import os
import sqlite3
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FILE_PATH  = "/mnt/c/Users/qendresa.krasniqi/Downloads/payments_statistics_datasets.xlsx"
OUTPUT_DIR = "/mnt/c/Users/qendresa.krasniqi/Downloads/final_summary"
os.makedirs(OUTPUT_DIR, exist_ok=True)

# --- Controlled vocabularies — mirrors Euro Area payments statistics framework ---
# Germany (DE) — largest euro area economy
VALID_COUNTRIES   = {"DE"}
VALID_INSTRUMENTS = {"credit_transfer","direct_debit","card_payment","e-money","cheque"}
VALID_TX_TYPES    = {"domestic","cross_border_intra_eu","cross_border_extra_eu"}
VALID_PSP_TYPES   = {"credit_institution","e-money_institution",
                     "payment_institution","post_office"}
VALID_CARD_TYPES  = {"debit","credit","prepaid"}
VALID_TERMINALS   = {"POS","ATM","online","contactless"}

# --- Excel styling helpers ---
def make_styles():
    header_font  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    header_fill  = PatternFill("solid", start_color="003366")
    data_font    = Font(name="Arial", size=10)
    alt_fill     = PatternFill("solid", start_color="EBF1F8")
    flag_fill    = PatternFill("solid", start_color="FFE0E0")
    thin         = Side(style="thin", color="CCCCCC")
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)
    return header_font, header_fill, data_font, alt_fill, flag_fill, border

def write_sheet(ws, df, styles, flag_col=None):
    """Write a DataFrame to an openpyxl worksheet with formatting."""
    header_font, header_fill, data_font, alt_fill, flag_fill, border = styles

    # Number formats applied by column name keyword
    INT_FORMAT   = '#,##0'          # e.g. 5,761,727,279
    FLOAT_FORMAT = '#,##0.00'       # e.g. 1,234.56
    PCT_FORMAT   = '#,##0.00'       # same for % columns — unit is already in header

    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border

    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        is_flagged = flag_col and getattr(row, flag_col.replace(" ","_"), 0) == 1
        fill = flag_fill if is_flagged else (alt_fill if row_idx % 2 == 0 else None)
        for col_idx, (col_name, value) in enumerate(zip(df.columns, row), 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font      = data_font
            cell.border    = border
            cell.alignment = Alignment(vertical="center")
            if fill:
                cell.fill = fill
            # Apply number format based on value type and column name
            # Columns that are counts — always whole numbers, no decimals
            whole_number_cols = {
                "Total Records", "Missing Records",
                "Imputed", "Dropped", "Duplicate Rows",
                "Outliers Detected", "Outlier Flag"
            }
            # Columns in millions — round to 2dp for consistent display across all sheets
            mn_cols = {
                "Volume (mn transactions)", "Total Transactions (mn)",
                "Card Transactions (mn)", "Raw Total (mn transactions)",
                "Pipeline Total (mn transactions)", "Difference (mn)"
            }
            if col_name in whole_number_cols or isinstance(value, (int, np.integer)):
                cell.number_format = INT_FORMAT
            elif isinstance(value, (float, np.floating)):
                cell.number_format = FLOAT_FORMAT

    for col_idx, col_name in enumerate(df.columns, 1):
        max_len = max(len(col_name), df.iloc[:, col_idx-1].astype(str).str.len().max())
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 40)
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"


# =============================================================================
# STEP 1 — INGEST
# Load quarterly payment datasets, schema validation, statistical summary
# =============================================================================

print("\n" + "="*65)
print("  STEP 1 — INGEST")
print("  Load quarterly payment datasets | Schema validation |")
print("  Statistical summary")
print("="*65)

# --- Load all sheets with exception handling ---
try:
    sheets = pd.read_excel(FILE_PATH, sheet_name=None)
except FileNotFoundError:
    raise FileNotFoundError(
        f"\n[ERROR] File not found: {FILE_PATH}"
        f"\nRun generate_data.py first to create the datasets."
    )
except Exception as e:
    raise RuntimeError(f"\n[ERROR] Failed to load workbook: {e}")

df_transactions = sheets["payments_transactions_quarterly"]
df_psps         = sheets["psp_reference_data"]
df_cards        = sheets["card_payments_detail"]

print(f"\n  [LOAD]")
print(f"  transactions : {df_transactions.shape[0]:>5,} rows × {df_transactions.shape[1]} columns")
print(f"  psps         : {df_psps.shape[0]:>5,} rows × {df_psps.shape[1]} columns")
print(f"  cards        : {df_cards.shape[0]:>5,} rows × {df_cards.shape[1]} columns")

# --- Inspect: head and tail of main dataset ---
print(f"\n  [INSPECT] transactions — first 3 rows:")
print(df_transactions.head(3).to_string())
print(f"\n  [INSPECT] transactions — last 3 rows:")
print(df_transactions.tail(3).to_string())

# --- Inspect: describe() for summary statistics ---
print(f"\n  [INSPECT] transactions — describe():")
print(df_transactions.describe().round(2).to_string())

# --- Null detection and imputation ---
# Strategy:
#   Numeric fields (total_value_eur_mn, number_of_transactions):
#     → impute with median of same country + instrument group
#     → rationale: series median is the best local estimate for a missing quarter
#   Key/identifier fields (reporting_country, quarter, payment_instrument):
#     → drop row — cannot impute a record identifier
print(f"\n  [NULL HANDLING]")

# Capture original completeness per country BEFORE imputation
# This reflects the raw data state — used in DQA sheet
raw_completeness_per_country = (
    df_transactions.groupby("reporting_country")
    .apply(lambda g: round(g.notnull().sum().sum() / g.size * 100, 2), include_groups=False)
    .rename("raw_completeness_pct")
)

impute_fields = ["total_value_eur_mn", "number_of_transactions"]
total_imputed = 0
total_dropped = 0

# Track imputed counts per country for DQA sheet
imputed_per_country = pd.Series(0, index=df_transactions["reporting_country"].unique(), name="imputed")
dropped_per_country = pd.Series(0, index=df_transactions["reporting_country"].unique(), name="dropped")

for field in impute_fields:
    null_mask  = df_transactions[field].isnull()
    null_count = null_mask.sum()
    if null_count > 0:
        # Record which country each null belongs to before imputing
        null_countries = df_transactions.loc[null_mask, "reporting_country"]
        imputed_per_country = imputed_per_country.add(
            null_countries.value_counts(), fill_value=0
        )
        # Impute with median of same country + instrument group
        df_transactions[field] = df_transactions.groupby(
            ["reporting_country", "payment_instrument"]
        )[field].transform(lambda x: x.fillna(x.median()))
        total_imputed += null_count
        print(f"  [IMPUTED] {field:<30} → {null_count} nulls filled with group median")
    else:
        print(f"  [OK]      {field:<30} → no nulls")

# Drop rows where key fields are null — cannot be imputed
key_fields = ["reporting_country", "quarter", "payment_instrument"]
key_nulls  = df_transactions[key_fields].isnull().any(axis=1)
total_dropped = key_nulls.sum()
if total_dropped > 0:
    dropped_countries = df_transactions.loc[key_nulls, "reporting_country"]
    dropped_per_country = dropped_per_country.add(
        dropped_countries.value_counts(), fill_value=0
    )
    df_transactions = df_transactions[~key_nulls].reset_index(drop=True)
    print(f"  [DROPPED] {total_dropped} rows with null key fields (not imputable)")
else:
    print(f"  [OK]      key fields (country, quarter, instrument) → no nulls")

print(f"\n  Summary: {total_imputed} values imputed | {total_dropped} rows dropped")

# --- Schema validation: check expected columns exist ---
print(f"\n  [SCHEMA VALIDATION]")
expected = {
    "transactions": ["reporting_country","quarter","payment_instrument",
                     "transaction_type","number_of_transactions","total_value_eur_mn"],
    "psps":         ["psp_id","psp_name","psp_type","home_country",
                     "license_date","is_active","psd2_licensed"],
    "cards":        ["reporting_country","period","card_type","transaction_type",
                     "at_terminal_type","number_of_cards_issued_mn",
                     "number_of_transactions","total_value_eur_mn"],
}
schema_ok = True
for name, cols in expected.items():
    df = {"transactions": df_transactions, "psps": df_psps, "cards": df_cards}[name]
    missing = [c for c in cols if c not in df.columns]
    if missing:
        print(f"  [FAIL] {name} missing columns: {missing}")
        schema_ok = False
    else:
        print(f"  [PASS] {name} schema valid")

# --- Statistical summary ---
print(f"\n  [STATISTICAL SUMMARY] transaction values (EUR mn)")
summary = df_transactions["total_value_eur_mn"].agg(
    ["min","max","mean","median","std"]
).round(2)
for stat, val in summary.items():
    print(f"  {stat:<8}: {val:>15,.2f}")

print(f"\n  By payment instrument (mean value EUR mn):")
by_inst = (
    df_transactions
    .groupby("payment_instrument")["total_value_eur_mn"]
    .agg(min="min", max="max", mean="mean", median="median")
    .round(2)
    .sort_values("mean", ascending=False)
)
print(by_inst.to_string())


# =============================================================================
# STEP 2 — VALIDATE
# Business rules | Referential integrity |
# Data quality assessment | Reconciliation checks
# =============================================================================

print("\n" + "="*65)
print("  STEP 2 — VALIDATE")
print("  Business rules | Referential integrity |")
print("  Data quality assessment")
print("="*65)

validation_log  = []   # stores all check results for TEST step report
validation_failures = []

def check(rule_id, description, passed, fail_count=0):
    """Log a validation check result."""
    status = "PASS" if passed else "FAIL"
    validation_log.append({
        "rule_id":     rule_id,
        "description": description,
        "status":      status,
        "fail_count":  fail_count,
    })
    icon = "✓" if passed else "✗"
    suffix = f" → {fail_count} rows" if not passed else ""
    print(f"  [{status}] {icon} {rule_id} — {description}{suffix}")
    if not passed:
        validation_failures.append({"rule_id": rule_id,
                                    "description": description,
                                    "fail_count": fail_count})

print(f"\n  Business Rules — Transactions:")
n = (~df_transactions["reporting_country"].isin(VALID_COUNTRIES)).sum()
check("VR001", "reporting_country valid EU code",        n == 0, n)

n = (~df_transactions["payment_instrument"].isin(VALID_INSTRUMENTS)).sum()
check("VR002", "payment_instrument from allowed list",   n == 0, n)

n = (~df_transactions["transaction_type"].isin(VALID_TX_TYPES)).sum()
check("VR003", "transaction_type from allowed list",     n == 0, n)

n = (df_transactions["number_of_transactions"] < 0).sum()
check("VR004", "number_of_transactions >= 0",            n == 0, n)

n = (df_transactions["total_value_eur_mn"] < 0).sum()
check("VR005", "total_value_eur_mn >= 0",                n == 0, n)

n = (~df_transactions["quarter"].str.match(r"^\d{4}Q[1-4]$", na=False)).sum()
check("VR006", "quarter format YYYYQn",                  n == 0, n)

print(f"\n  Referential Integrity — PSPs:")
n = (~df_psps["home_country"].isin(VALID_COUNTRIES)).sum()
check("VR007", "home_country valid EU code",             n == 0, n)

n = df_psps.duplicated(subset="psp_id").sum()
check("VR008", "psp_id unique",                          n == 0, n)

n = (~df_psps["psp_type"].isin(VALID_PSP_TYPES)).sum()
check("VR009", "psp_type from allowed list",             n == 0, n)

n = ((df_psps["is_active"] == True) & df_psps["license_date"].isnull()).sum()
check("VR010", "active PSPs have valid license date",    n == 0, n)

print(f"\n  Business Rules — Cards:")
n = (~df_cards["card_type"].isin(VALID_CARD_TYPES)).sum()
check("VR011", "card_type from allowed list",            n == 0, n)

n = (~df_cards["at_terminal_type"].isin(VALID_TERMINALS)).sum()
check("VR012", "at_terminal_type from allowed list",     n == 0, n)

n = (~df_cards["number_of_cards_issued_mn"].between(0, 500)).sum()
check("VR013", "cards_issued_mn between 0 and 500",      n == 0, n)

# --- Data Quality Assessment Report ---

# 1. Completeness — missing values across all columns and datasets
total_nulls = (
    df_transactions.isnull().sum().sum() +
    df_psps.isnull().sum().sum() +
    df_cards.isnull().sum().sum()
)
total_dupes = (
    df_transactions.duplicated().sum() +
    df_psps.duplicated().sum() +
    df_cards.duplicated().sum()
)
missing_records = total_nulls + total_dupes
completeness_pass = missing_records == 0

# 2. Validity — invalid payment values (negatives) and invalid codes
invalid_values = (
    (df_transactions["total_value_eur_mn"] < 0).sum() +
    (df_transactions["number_of_transactions"] < 0).sum() +
    (df_cards["total_value_eur_mn"] < 0).sum()
)
validity_pass = invalid_values == 0

# 3. Consistency — country totals reconcile across datasets
tx_countries  = set(df_transactions["reporting_country"].unique())
card_countries = set(df_cards["reporting_country"].unique())
countries_match = tx_countries == card_countries
missing_countries = VALID_COUNTRIES - tx_countries
consistency_pass = countries_match and len(missing_countries) == 0

# 4. Statistical quality — unusual movements (outlier detection preview)
# Quick z-score pass to count suspicious series before full outlier step
zscore_preview = (
    df_transactions
    .groupby(["reporting_country","payment_instrument"])["total_value_eur_mn"]
    .transform(lambda x: (x - x.mean()) / x.std())
)
unusual_movements = (zscore_preview.abs() > 3).sum()
review_required = unusual_movements > 0

# --- Print DQA Report ---
dqa_lines = [
    f"\n  Data Quality Assessment Report",
    f"  {'='*40}",
    f"",
    f"  1. Completeness",
    f"     Missing records  : {missing_records}",
    f"     Imputed          : {total_imputed}",
    f"     Dropped          : {total_dropped}",
    f"     Status           : {'PASS' if completeness_pass else 'FAIL'} (after imputation)",
    f"",
    f"  2. Validity",
    f"     Invalid payment values : {invalid_values}",
    f"     Status                 : {'PASS' if validity_pass else 'FAIL'}",
    f"",
    f"  3. Consistency",
    f"     Country totals reconcile  : {'PASS' if consistency_pass else 'FAIL'}",
    f"     Missing expected countries: {len(missing_countries)}",
    f"",
    f"  4. Statistical Quality",
    f"     Unusual movements detected : {unusual_movements}",
    f"     Review required            : {'YES' if review_required else 'NO'}",
    f"  {'='*40}",
]
dqa_report = "\n".join(dqa_lines)
print(dqa_report)

# Log into validation_log for TEST step
check("DQA-01", "Completeness — no missing or duplicate records", completeness_pass, missing_records)
check("DQA-02", "Validity — no invalid payment values",           validity_pass,     invalid_values)
check("DQA-03", "Consistency — country totals reconcile",         consistency_pass,  len(missing_countries))
check("DQA-04", "Statistical quality — unusual movements flagged",True, unusual_movements)

total_failures = sum(f["fail_count"] for f in validation_failures)
print(f"\n  Validation complete — {len(validation_failures)} rules failed, "
      f"{total_failures} total rows flagged")


# =============================================================================
# STEP 3 — TRANSFORM
# Data joins | Derived variables |
# Statistical indicators | QoQ and YoY calculations
# =============================================================================

print("\n" + "="*65)
print("  STEP 3 — TRANSFORM")
print("  Data joins | Derived variables |")
print("  Statistical indicators | QoQ and YoY calculations")
print("="*65)

# --- Sort chronologically — critical for time-series calculations ---
df_transactions = df_transactions.sort_values(
    ["reporting_country", "payment_instrument", "quarter"]
).reset_index(drop=True)

# --- Derived variable: average transaction value (EUR) ---
# Computed here as a derived metric — not a reported field from PSPs
df_transactions["avg_value_eur"] = (
    df_transactions["total_value_eur_mn"] * 1_000_000
    / df_transactions["number_of_transactions"]
).round(2)
print(f"\n  [DERIVED] avg_value_eur = total_value_eur_mn × 1M / number_of_transactions")

# --- QoQ (Quarter-on-Quarter) % change ---
# Compares each quarter to the immediately preceding quarter
# within the same country + instrument group
# First quarter per group will be NaN — correct and expected
df_transactions["qoq_change_pct"] = (
    df_transactions
    .groupby(["reporting_country", "payment_instrument", "transaction_type"])["total_value_eur_mn"]
    .pct_change() * 100
).round(2)
print(f"  [DERIVED] qoq_change_pct = % change vs previous quarter")

# --- YoY (Year-on-Year) % change ---
# Compares each quarter to the same quarter one year prior (4 quarters back)
# e.g. 2024-Q2 vs 2023-Q2 — removes seasonal effects unlike QoQ
# First 4 quarters per group will be NaN — correct and expected
df_transactions["yoy_change_pct"] = (
    df_transactions
    .groupby(["reporting_country", "payment_instrument", "transaction_type"])["total_value_eur_mn"]
    .pct_change(periods=4) * 100
).round(2)
print(f"  [DERIVED] yoy_change_pct = % change vs same quarter one year prior")

# --- Cards aggregated to country level (used later in Step 5) ---
cards_by_country = (
    df_cards
    .groupby("reporting_country")
    .agg(
        card_transactions_mn = ("number_of_transactions", lambda x: x.sum() / 1_000_000),
        card_value_eur_mn    = ("total_value_eur_mn",     "sum"),
    )
    .reset_index()
)

# Merge row count validation
n_tx_countries   = df_transactions["reporting_country"].nunique()
n_card_countries = df_cards["reporting_country"].nunique()
print(f"\n  [JOIN VALIDATION]")
print(f"  Unique countries in transactions : {n_tx_countries}")
print(f"  Unique countries in cards        : {n_card_countries}")

# Dominant instrument per country (by volume)
dominant = (
    df_transactions
    .groupby(["reporting_country","payment_instrument"])["number_of_transactions"]
    .sum().reset_index()
    .sort_values("number_of_transactions", ascending=False)
    .groupby("reporting_country").first().reset_index()
    [["reporting_country","payment_instrument"]]
    .rename(columns={"payment_instrument": "dominant_instrument"})
)

# Cross-border share per country
df_transactions["is_cross_border"] = (
    df_transactions["transaction_type"].str.startswith("cross_border")
)
cross_border = (
    df_transactions.groupby("reporting_country")["is_cross_border"]
    .mean().mul(100).round(1).reset_index()
    .rename(columns={"is_cross_border": "cross_border_share_pct"})
)
print(f"  [INDICATORS] dominant_instrument and cross_border_share_pct computed")


# =============================================================================
# STEP 4 — ANALYSE
# Country-level trends | Outlier detection (Z-score backend only) |
# Quarter-on-quarter spike detection
# =============================================================================

print("\n" + "="*65)
print("  STEP 4 — ANALYSE")
print("  Country-level trends | Outlier detection (Z-score) |")
print("  Quarter-on-quarter spike detection")
print("="*65)

# --- Backend: Z-score calculated per country + instrument series ---
# Z-score value is never exposed in the output — used only to derive the flag
df_transactions["_zscore"] = (
    df_transactions
    .groupby(["reporting_country","payment_instrument","transaction_type"])["total_value_eur_mn"]
    .transform(lambda x: (x - x.mean()) / x.std())
)

# Outlier Flag: 1 if z-score exceeds threshold, 0 otherwise
df_transactions["outlier_flag"] = (
    df_transactions["_zscore"].abs() > 3
).astype(int)

# Drop internal z-score — backend only, never leaves the pipeline
df_transactions = df_transactions.drop(columns=["_zscore"])

total_outliers = df_transactions["outlier_flag"].sum()
print(f"\n  [OUTLIER]  flagged {total_outliers} rows via z-score (|z| > 3)")
print(f"  [NOTE]     z-score computed internally — not exposed in output")

# --- IQR outlier detection — used for detailed outlier report ---
# Identifies values beyond 1.5 x IQR from Q1/Q3 per instrument + transaction type
# This is what the boxplot visualises
def detect_iqr_outliers(group):
    Q1  = group["total_value_eur_mn"].quantile(0.25)
    Q3  = group["total_value_eur_mn"].quantile(0.75)
    IQR = Q3 - Q1
    mask = (
        (group["total_value_eur_mn"] < Q1 - 1.5 * IQR) |
        (group["total_value_eur_mn"] > Q3 + 1.5 * IQR)
    )
    result = group[mask].copy()
    result["Q1"]            = round(Q1, 2)
    result["Q3"]            = round(Q3, 2)
    result["IQR"]           = round(IQR, 2)
    result["lower_bound"]   = round(Q1 - 1.5 * IQR, 2)
    result["upper_bound"]   = round(Q3 + 1.5 * IQR, 2)
    return result

# Apply IQR detection — use include_groups=False to preserve group keys in result
outlier_list = []
for (instrument, txn_type), group in df_transactions.groupby(["payment_instrument","transaction_type"]):
    result = detect_iqr_outliers(group)
    outlier_list.append(result)

df_iqr_outliers = pd.concat(outlier_list, ignore_index=True) if outlier_list else pd.DataFrame()

# Add value in EUR bn for readability
df_iqr_outliers["value_eur_bn"]       = (df_iqr_outliers["total_value_eur_mn"] / 1000).round(2)
df_iqr_outliers["lower_bound_eur_bn"] = (df_iqr_outliers["lower_bound"] / 1000).round(2)
df_iqr_outliers["upper_bound_eur_bn"] = (df_iqr_outliers["upper_bound"] / 1000).round(2)

print(f"  [IQR]      flagged {len(df_iqr_outliers)} rows (beyond 1.5×IQR)")

if len(df_iqr_outliers) > 0:
    print(f"\n  Outlier detail (IQR method):")
    display_cols = ["quarter","payment_instrument","transaction_type",
                    "value_eur_bn","lower_bound_eur_bn","upper_bound_eur_bn"]
    print(df_iqr_outliers[display_cols].to_string(index=False))

# --- Country-level trend summary ---
print(f"\n  [TRENDS] QoQ and YoY by instrument (mean across transaction types):")
trends = (
    df_transactions
    .groupby("payment_instrument")[["qoq_change_pct","yoy_change_pct"]]
    .mean().round(2)
    .sort_values("yoy_change_pct", ascending=False)
)
print(trends.to_string())


# =============================================================================
# STEP 5 — REPORT
# Germany payments statistical summary tables | Dashboard-ready outputs
# Sheets: Payments Statistics | Data Quality Assessment | Country Indicators | Metadata
# =============================================================================

print("\n" + "="*65)
print("  STEP 5 — REPORT")
print("  Germany payments statistical summary tables | Dashboard-ready outputs")
print("="*65)

styles = make_styles()

# --- Main report: country × instrument × quarter ---
# Includes QoQ, YoY and Outlier Flag as derived columns
report = (
    df_transactions
    .groupby(["reporting_country","payment_instrument","quarter"])
    .agg(
        volume_mn      = ("number_of_transactions", lambda x: x.sum() / 1_000_000),
        value_eur_bn   = ("total_value_eur_mn",     lambda x: round(x.sum() / 1_000,     2)),
        qoq_change_pct = ("qoq_change_pct",         "first"),
        yoy_change_pct = ("yoy_change_pct",          "first"),
        outlier_flag   = ("outlier_flag",            "max"),
    )
    .reset_index()
)

report["avg_value_eur"] = (
    report["value_eur_bn"] * 1_000_000_000
    / (report["volume_mn"] * 1_000_000)
).round(2)

# Format period label e.g. 2024-Q1
report["period"] = report["quarter"].str.replace(
    r"(\d{4})Q(\d)", r"\1-Q\2", regex=True
)

report = report.sort_values(
    ["reporting_country","payment_instrument","quarter"]
).reset_index(drop=True).drop(columns=["quarter"])

report.columns = [
    "Country", "Payment Instrument",
    "Volume (mn transactions)", "Value (EUR bn)",
    "QoQ Change (%)", "YoY Change (%)",
    "Outlier Flag", "Avg Value per Transaction (EUR)", "Period"
]

# Reorder columns for readability
report = report[[
    "Country", "Payment Instrument", "Period",
    "Volume (mn transactions)", "Value (EUR bn)",
    "Avg Value per Transaction (EUR)",
    "QoQ Change (%)", "YoY Change (%)", "Outlier Flag"
]]

print(f"\n  Main report : {len(report):,} rows (country × instrument × quarter)")
print(f"  Outliers    : {report['Outlier Flag'].sum()} flagged rows")

# Derive country totals from report — single source of truth for all sheets
tx_by_country = (
    report
    .groupby("Country")
    .agg(
        total_transactions_mn = ("Volume (mn transactions)", lambda x: round(x.sum(), 2)),
        total_value_eur_mn    = ("Value (EUR bn)",           lambda x: round(x.sum() * 1_000, 2)),
    )
    .reset_index()
    .rename(columns={"Country": "reporting_country"})
)
country_combined = tx_by_country.merge(cards_by_country, on="reporting_country", how="left")

# Validate join — check for unmatched records
unmatched = country_combined["card_transactions_mn"].isnull().sum()
print(f"  Countries before join : {len(tx_by_country)}")
print(f"  Countries after join  : {len(country_combined)}")
print(f"  Unmatched records     : {unmatched} "
      f"{'← WARNING' if unmatched > 0 else '← OK'}")

# Statistical indicators per country
country_combined["card_share_pct"] = (
    country_combined["card_transactions_mn"]
    / country_combined["total_transactions_mn"] * 100
).round(1)

country_combined["avg_tx_value_eur"] = (
    country_combined["total_value_eur_mn"] * 1_000_000
    / (country_combined["total_transactions_mn"] * 1_000_000)
).round(2)

indicators = (
    country_combined
    .merge(dominant,     on="reporting_country", how="left")
    .merge(cross_border, on="reporting_country", how="left")
    .sort_values("total_value_eur_mn", ascending=False)
)
print(f"  [JOIN] transactions + cards on reporting_country → {len(country_combined)} countries")
print(f"  [INDICATORS] total_transactions_mn matches Volume (mn transactions) in Payments Statistics sheet")

# --- Data Quality Assessment sheet — country level ---
dqa_country = (
    df_transactions
    .groupby("reporting_country")
    .agg(
        total_rows          = ("total_value_eur_mn",    "count"),
        missing_values      = ("total_value_eur_mn",    lambda x: x.isnull().sum()),
        negative_values     = ("total_value_eur_mn",    lambda x: (x < 0).sum()),
        duplicate_rows      = ("total_value_eur_mn",    lambda x: x.duplicated().sum()),
        outliers_detected   = ("outlier_flag",          "sum"),
        avg_qoq_change_pct  = ("qoq_change_pct",        "mean"),
        avg_yoy_change_pct  = ("yoy_change_pct",        "mean"),
    )
    .reset_index()
)

# Derived quality indicators per country
# Completeness % uses RAW data before imputation — honest representation of source quality
dqa_country["completeness_pct"] = dqa_country["reporting_country"].map(
    raw_completeness_per_country
).round(2)

dqa_country["validity_status"]      = dqa_country["negative_values"].apply(
    lambda x: "PASS" if x == 0 else "FAIL"
)
dqa_country["completeness_status"]  = dqa_country["missing_values"].apply(
    lambda x: "PASS" if x == 0 else "FAIL"
)
dqa_country["review_required"]      = dqa_country["outliers_detected"].apply(
    lambda x: "YES" if x > 0 else "NO"
)
dqa_country["avg_qoq_change_pct"]   = dqa_country["avg_qoq_change_pct"].round(2)
dqa_country["avg_yoy_change_pct"]   = dqa_country["avg_yoy_change_pct"].round(2)

dqa_country = dqa_country.rename(columns={
    "reporting_country":    "Country",
    "total_rows":           "Total Records",
    "missing_values":       "Missing Values",
    "negative_values":      "Invalid Values",
    "duplicate_rows":       "Duplicate Rows",
    "outliers_detected":    "Outliers Detected",
    "avg_qoq_change_pct":   "Avg QoQ Change (%)",
    "avg_yoy_change_pct":   "Avg YoY Change (%)",
    "completeness_pct":     "Completeness (%)",
    "validity_status":      "Validity",
    "completeness_status":  "Completeness",
    "review_required":      "Review Required",
})[["Country","Total Records","Missing Values","Completeness (%)","Completeness",
    "Invalid Values","Validity","Duplicate Rows",
    "Outliers Detected","Review Required",
    "Avg QoQ Change (%)","Avg YoY Change (%)"]]

dqa_country = dqa_country.sort_values("Country").reset_index(drop=True)

# Add missing records, imputed and dropped counts per country
dqa_country["Missing Records"] = dqa_country["Country"].map(
    imputed_per_country.add(dropped_per_country, fill_value=0)
).fillna(0).astype(int)

dqa_country["Imputed"] = dqa_country["Country"].map(
    imputed_per_country
).fillna(0).astype(int)

dqa_country["Dropped"] = dqa_country["Country"].map(
    dropped_per_country
).fillna(0).astype(int)

# Reorder to put new columns near the front
dqa_country = dqa_country[[
    "Country", "Total Records",
    "Missing Records", "Imputed", "Dropped",
    "Completeness (%)", "Completeness",
    "Invalid Values", "Validity", "Duplicate Rows",
    "Outliers Detected", "Review Required",
    "Avg QoQ Change (%)", "Avg YoY Change (%)"
]]

# --- Build and save Excel workbook ---
wb = Workbook()

# Sheet 1: Payments Statistics — main report with outlier flag highlighted
ws_main = wb.active
ws_main.title = "Payments Statistics"
write_sheet(ws_main, report, styles, flag_col="Outlier_Flag")

# Sheet 2: Data Quality Assessment — country level
ws_dqa = wb.create_sheet("Data Quality Assessment")
write_sheet(ws_dqa, dqa_country, styles)

# Sheet 3: Outlier Report — IQR method detailed findings
ws_outliers = wb.create_sheet("Outlier Report")
# Note: Outlier Report covers ALL transaction types (domestic + cross-border)
# The boxplot chart (outlier_boxplot.R) filters to domestic transactions only
# for visual clarity. Both use the same IQR method (1.5 × IQR from Q1/Q3).
if len(df_iqr_outliers) > 0:
    outlier_report = df_iqr_outliers[[
        "quarter", "payment_instrument", "transaction_type",
        "value_eur_bn", "lower_bound_eur_bn", "upper_bound_eur_bn",
        "Q1", "Q3", "IQR"
    ]].rename(columns={
        "quarter":              "Quarter",
        "payment_instrument":   "Payment Instrument",
        "transaction_type":     "Transaction Type",
        "value_eur_bn":         "Value (EUR bn)",
        "lower_bound_eur_bn":   "Lower Bound (EUR bn)",
        "upper_bound_eur_bn":   "Upper Bound (EUR bn)",
        "Q1":                   "Q1 (EUR mn)",
        "Q3":                   "Q3 (EUR mn)",
        "IQR":                  "IQR (EUR mn)",
    }).sort_values(["Payment Instrument","Quarter"]).reset_index(drop=True)
    write_sheet(ws_outliers, outlier_report, styles)
    # Add note below the table
    note_row = len(outlier_report) + 3
    ws_outliers.cell(row=note_row, column=1,
        value="Note: This report covers all transaction types (domestic + cross-border). "
              "The boxplot chart filters to domestic transactions only for visual clarity. "
              "Both use the IQR method (beyond 1.5 × IQR from Q1/Q3)."
    ).font = Font(name="Arial", size=9, italic=True, color="666666")
    print(f"  Outlier Report sheet: {len(outlier_report)} flagged rows")
else:
    ws_outliers.cell(row=1, column=1, value="No outliers detected by IQR method.")

# Sheet 4: Country Indicators
ws_indicators = wb.create_sheet("Country Indicators")
indicators_export = indicators.rename(columns={
    "reporting_country":      "Country",
    "total_transactions_mn":  "Total Transactions (mn)",
    "total_value_eur_mn":     "Total Value (EUR mn)",
    "card_transactions_mn":   "Card Transactions (mn)",
    "card_value_eur_mn":      "Card Value (EUR mn)",
    "card_share_pct":         "Card Share (%)",
    "avg_tx_value_eur":       "Avg Tx Value (EUR)",
    "dominant_instrument":    "Dominant Instrument",
    "cross_border_share_pct": "Cross-Border Share (%)",
})
write_sheet(ws_indicators, indicators_export, styles)

# Sheet 4: Metadata
ws_meta = wb.create_sheet("Metadata")
meta_rows = [
    ("Report Title",     "Germany Payments Statistics — Quarterly Series"),
    ("Reference Period", "2019-Q1 to 2024-Q4"),
    ("Generated",        datetime.now().strftime("%Y-%m-%d %H:%M")),
    ("Data Source",      "Synthetic data — Euro Area payments statistics framework"),
    ("Countries",        ", ".join(sorted(report["Country"].unique()))),
    ("Country Coverage", "Germany (DE) — largest euro area economy"),
    ("Instruments",      ", ".join(sorted(report["Payment Instrument"].unique()))),
    ("Total Rows",       len(report)),
    ("Outliers Flagged", int(report["Outlier Flag"].sum())),
    ("Outlier Flag",     "1 = statistically unusual observation, 0 = normal. Identified using z-score method per country and instrument series."),
    ("Rounding Note",    "Figures are stored at full precision and displayed rounded to 2 decimal places. Country-level totals in Country Indicators are rounded once after aggregation. Individual rows in Payments Statistics retain full precision — sums may therefore not add up exactly due to rounding, consistent with Euro Area statistical publication practice."),
    ("YoY definition",   "% change vs same quarter one year prior (4 quarters back)"),
]
bold   = Font(bold=True, name="Arial")
normal = Font(name="Arial")
for r, (key, val) in enumerate(meta_rows, 1):
    ws_meta.cell(row=r, column=1, value=key).font  = bold
    ws_meta.cell(row=r, column=2, value=str(val)).font = normal
ws_meta.column_dimensions["A"].width = 22
ws_meta.column_dimensions["B"].width = 70

try:
    wb.save(f"{OUTPUT_DIR}/payments_quarterly_report.xlsx")
    print(f"\n  Saved → payments_quarterly_report.xlsx")
    print(f"  Sheets: Payments Statistics | Data Quality Assessment | Country Indicators | Metadata")
except PermissionError:
    raise PermissionError(
        f"\n[ERROR] Cannot save — file may be open in Excel. "
        f"Please close it and run the pipeline again."
    )
except Exception as e:
    raise RuntimeError(f"\n[ERROR] Failed to save workbook: {e}")


# =============================================================================
# STEP 6 — SQL
# Business queries executed via sqlite3 against the transformed dataset.
# Results saved as separate Excel reports per business question.
#
# Query 1: Which countries and instruments have the highest average
#          transaction value? — helps identify high-value payment corridors
# Query 2: Which countries showed the strongest YoY growth in 2024?
#          — identifies emerging payment trends
# =============================================================================

print("\n" + "="*65)
print("  STEP 6 — SQL")
print("  Business queries | sqlite3 in-memory database |")
print("  Results saved as separate Excel reports")
print("="*65)

# --- Load transformed transactions into an in-memory SQLite database ---
# sqlite3 is built into Python — no external database needed
conn = sqlite3.connect(":memory:")

df_transactions.to_sql("payments_transactions", conn, index=False, if_exists="replace")
print(f"\n  [SQL] Loaded {len(df_transactions):,} rows into in-memory SQLite database")

# =============================================================================
# QUERY — Instrument Growth 2019 vs 2024
# Business question: Which payment instrument grew the most between 2019 and 2024?
# Pure SQL using CTEs — no pre-computed Python columns used.
# All growth calculations done directly inside the query.
# =============================================================================

query3 = """
    WITH base_year AS (
        SELECT
            payment_instrument,
            ROUND(SUM(total_value_eur_mn) / 1000.0, 2)        AS value_eur_bn_2019,
            ROUND(SUM(number_of_transactions) / 1000000.0, 2)  AS volume_mn_2019
        FROM payments_transactions
        WHERE quarter LIKE '2019%'
          AND reporting_country = 'DE'
        GROUP BY payment_instrument
    ),
    end_year AS (
        SELECT
            payment_instrument,
            ROUND(SUM(total_value_eur_mn) / 1000.0, 2)        AS value_eur_bn_2024,
            ROUND(SUM(number_of_transactions) / 1000000.0, 2)  AS volume_mn_2024
        FROM payments_transactions
        WHERE quarter LIKE '2024%'
          AND reporting_country = 'DE'
        GROUP BY payment_instrument
    )
    SELECT
        b.payment_instrument                                          AS [Payment Instrument],
        b.value_eur_bn_2019                                           AS [Value 2019 (EUR bn)],
        e.value_eur_bn_2024                                           AS [Value 2024 (EUR bn)],
        ROUND((e.value_eur_bn_2024 - b.value_eur_bn_2019)
              / b.value_eur_bn_2019 * 100, 2)                         AS [Value Growth (%)],
        b.volume_mn_2019                                              AS [Volume 2019 (mn)],
        e.volume_mn_2024                                              AS [Volume 2024 (mn)],
        ROUND((e.volume_mn_2024 - b.volume_mn_2019)
              / b.volume_mn_2019 * 100, 2)                            AS [Volume Growth (%)]
    FROM base_year b
    JOIN end_year e ON b.payment_instrument = e.payment_instrument
    ORDER BY [Value Growth (%)] DESC
"""

df_q3 = pd.read_sql_query(query3, conn)
print(f"\n  [QUERY] Instrument growth 2019 vs 2024 (pure SQL — no pre-computed columns)")
print(f"  Business question: Which payment instrument grew the most between 2019 and 2024?")
print(df_q3.to_string(index=False))

# Add query result as sheet to the main report
wb_main = __import__("openpyxl").load_workbook(f"{OUTPUT_DIR}/payments_quarterly_report.xlsx")
ws_q3 = wb_main.create_sheet("Instrument Growth 2019-2024")
write_sheet(ws_q3, df_q3, styles)
wb_main.save(f"{OUTPUT_DIR}/payments_quarterly_report.xlsx")
print(f"  Saved → Instrument Growth 2019-2024 sheet added to payments_quarterly_report.xlsx")

conn.close()
print(f"\n  [SQL] Database connection closed")


# =============================================================================
# STEP 7 — TEST
# Unit tests | Pipeline validation |
# Output consistency checks and validation report
# =============================================================================

print("\n" + "="*65)
print("  STEP 7 — TEST")
print("  Unit tests | Pipeline validation |")
print("  Reconciliation checks | Output consistency checks")
print("="*65)

test_results = []

def test(name, passed, detail=""):
    status = "PASS" if passed else "FAIL"
    test_results.append({"test": name, "status": status, "detail": detail})
    icon = "✓" if passed else "✗"
    suffix = f" ({detail})" if detail else ""
    print(f"  [{status}] {icon} {name}{suffix}")

print(f"\n  Unit Tests:")
# Check derived fields are non-negative
test("avg_value_eur >= 0",
     (df_transactions["avg_value_eur"] >= 0).all())

# Check QoQ is null only for first quarter per group
first_quarters = df_transactions.groupby(
    ["reporting_country","payment_instrument","transaction_type"])["quarter"].transform("min")
expected_nulls = (df_transactions["quarter"] == first_quarters).sum()
actual_nulls   = df_transactions["qoq_change_pct"].isna().sum()
test("QoQ null only for first quarter per group",
     actual_nulls == expected_nulls,
     f"expected {expected_nulls} nulls, got {actual_nulls}")

# Check YoY is null only for first 4 quarters per group
actual_yoy_nulls = df_transactions["yoy_change_pct"].isna().sum()
expected_yoy_nulls = (
    df_transactions
    .groupby(["reporting_country","payment_instrument","transaction_type"])["quarter"]
    .transform(lambda x: (x.rank(method="first") <= 4))
    .sum()
)
test("YoY null only for first 4 quarters per group",
     actual_yoy_nulls == expected_yoy_nulls,
     f"expected {int(expected_yoy_nulls)} nulls, got {actual_yoy_nulls}")

print(f"\n  Pipeline Validation:")
test("Report has expected row count",
     len(report) == df_transactions["quarter"].nunique() * df_transactions["payment_instrument"].nunique(),
     f"{len(report):,} rows")

test("Expected country present in report",
     set(report["Country"]) == {"DE"})

test("All instruments present in report",
     set(report["Payment Instrument"]) == VALID_INSTRUMENTS)

test("No nulls in report main columns",
     report[["Country","Payment Instrument","Period",
             "Volume (mn transactions)","Value (EUR bn)"]].isnull().sum().sum() == 0)

print(f"\n  Output Consistency Checks:")
report_total_bn  = report["Value (EUR bn)"].sum()
report_total_vol = report["Volume (mn transactions)"].sum()
test("Report totals are positive and non-zero",
     report_total_bn > 0 and report_total_vol > 0,
     f"value={report_total_bn:,.0f} EUR bn, volume={report_total_vol:,.0f} mn")

test("Outlier flag is binary (0 or 1)",
     set(df_transactions["outlier_flag"].unique()).issubset({0,1}))

# Check for impossible growth rates — flags extreme values for review
# Note: synthetic data includes ~2% injected outliers which produce large spikes
impossible_qoq = (df_transactions["qoq_change_pct"].abs() > 10000).sum()
impossible_yoy = (df_transactions["yoy_change_pct"].abs() > 10000).sum()
test("No impossible QoQ growth rates (> ±10000%)", impossible_qoq == 0,
     f"{impossible_qoq} rows flagged for review" if impossible_qoq > 0 else "")
test("No impossible YoY growth rates (> ±10000%)", impossible_yoy == 0,
     f"{impossible_yoy} rows flagged for review" if impossible_yoy > 0 else "")

# Check quarter date range is within expected period
quarters = df_transactions["quarter"].unique()
years = [int(q[:4]) for q in quarters]
test("All quarters within expected period (2019-2024)",
     min(years) >= 2019 and max(years) <= 2024,
     f"years found: {min(years)}–{max(years)}")

test("Validation log complete",
     len(validation_log) > 0,
     f"{len(validation_log)} checks logged")

# --- Reconciliation: raw file vs pipeline output per country ---
# Source of truth: raw Excel file (as reported by national central banks)
# Pipeline output: Total Transactions in Country Indicators
print(f"\n  Reconciliation — Raw File vs Pipeline Output:")

raw_file = sheets["payments_transactions_quarterly"]

raw_totals = (
    raw_file
    .groupby("reporting_country")["number_of_transactions"]
    .sum()
    .div(1_000_000)
    .reset_index()
    .rename(columns={
        "reporting_country":      "Country",
        "number_of_transactions": "Raw Total (mn transactions)",
    })
)

pipeline_totals = (
    indicators[["reporting_country","total_transactions_mn"]]
    .rename(columns={
        "reporting_country":      "Country",
        "total_transactions_mn":  "Pipeline Total (mn transactions)",
    })
)

recon_df = raw_totals.merge(pipeline_totals, on="Country", how="outer")
recon_df["Difference (mn)"] = (
    recon_df["Pipeline Total (mn transactions)"] - recon_df["Raw Total (mn transactions)"]
).round(3)
recon_df["Status"] = recon_df["Difference (mn)"].abs().apply(
    lambda x: "PASS" if x < 0.1 else "FAIL"
)
recon_df = recon_df.sort_values("Country").reset_index(drop=True)

# Summary row
summary_row = pd.DataFrame([{
    "Country":                          "TOTAL",
    "Raw Total (mn transactions)":       recon_df["Raw Total (mn transactions)"].sum(),
    "Pipeline Total (mn transactions)":  recon_df["Pipeline Total (mn transactions)"].sum(),
    "Difference (mn)":                   recon_df["Difference (mn)"].sum(),
    "Status":                           "PASS" if (recon_df["Status"] == "PASS").all() else "FAIL",
}])
recon_df = pd.concat([recon_df, summary_row], ignore_index=True)

recon_pass_all = (recon_df[recon_df["Country"] != "TOTAL"]["Status"] == "PASS").all()
n_pass = (recon_df[recon_df["Country"] != "TOTAL"]["Status"] == "PASS").sum()
n_total = len(recon_df) - 1

test("Reconciliation — raw vs pipeline totals match",
     recon_pass_all,
     f"{n_pass}/{n_total} countries matched")

# Add Reconciliation sheet to existing workbook
wb2 = __import__("openpyxl").load_workbook(f"{OUTPUT_DIR}/payments_quarterly_report.xlsx")
ws_recon = wb2.create_sheet("Reconciliation")
write_sheet(ws_recon, recon_df, styles)
wb2.save(f"{OUTPUT_DIR}/payments_quarterly_report.xlsx")
print(f"  Reconciliation sheet added to workbook")

# --- Validation Report ---
passes   = sum(1 for v in validation_log if v["status"] == "PASS")
fails    = sum(1 for v in validation_log if v["status"] == "FAIL")
t_passes = sum(1 for t in test_results    if t["status"] == "PASS")
t_fails  = sum(1 for t in test_results    if t["status"] == "FAIL")

null_count  = sheets["payments_transactions_quarterly"].isnull().sum().sum()
dupe_count  = df_transactions.duplicated().sum()
invalid_cc  = (~df_transactions["reporting_country"].isin(VALID_COUNTRIES)).sum()
neg_values  = (df_transactions["total_value_eur_mn"] < 0).sum()
quarter_ok  = (~df_transactions["quarter"].str.match(r"^\d{4}Q[1-4]$", na=False)).sum() == 0
recon_pass  = recon_pass_all

report_lines = [
    "",
    "=" * 45,
    "  VALIDATION REPORT",
    f"  Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
    "─" * 45,
    f"  {'✓' if null_count  == 0 else '✗'} Missing values        : {null_count}",
    f"  {'✓' if dupe_count  == 0 else '✗'} Duplicate records     : {dupe_count}",
    f"  {'✓' if invalid_cc  == 0 else '✗'} Invalid country codes : {invalid_cc}",
    f"  {'✓' if neg_values  == 0 else '✗'} Negative payment values: {neg_values}",
    f"  {'✓' if quarter_ok       else '✗'} Quarter sequence valid",
    f"  {'✓' if recon_pass       else '✗'} Reconciliation passed",
    f"  {'✓' if total_outliers>0 else '✗'} Outliers detected     : {total_outliers}",
    "─" * 45,
    f"  Business rules : {passes} passed / {fails} failed",
    f"  Unit tests     : {t_passes} passed / {t_fails} failed",
    "=" * 45,
    "",
]
validation_report = "\n".join(report_lines)
print(validation_report)

print("=" * 65)
print(f"  PIPELINE COMPLETE")
print(f"  Outputs saved to: {OUTPUT_DIR}")
print(f"  Files generated:")
print(f"    • payments_quarterly_report.xlsx")
print(f"      └─ Payments Statistics        ← quarterly report with QoQ, YoY, Outlier Flag")
print(f"      └─ Data Quality Assessment    ← country-level DQA report")
print(f"      └─ Country Indicators         ← country-level summary")
print(f"      └─ Reconciliation             ← raw vs pipeline totals by country")
print(f"      └─ Instrument Growth 2019-2024← SQL: which instrument grew most?")
print(f"      └─ Metadata                   ← definitions and run info")
print("=" * 65 + "\n")